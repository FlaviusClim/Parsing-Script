"""This is a script for parsing through CSV file for specific information"""
import os

import pandas as pd
import openpyxl


def get_user_input():
    """get_user_input function will take the number and the name of signals from the user"""

    nb_of_signals = int(input("How many signals do you need to search for? "))
    signal_names = []

    for i in range(nb_of_signals):
        signal = input(f"Give the signal {i + 1} here: ")
        signal_names.append(signal)

    return signal_names


def read_csv():
    """read_csv function will read the csv file for parsing and
    all the necessary datas will be stored in df variable"""

    df = pd.read_csv("CAN_Signal_Log.csv")

    return df


def search_signal(df, signal_name):
    """this function will parse through the df and search for specific signal
    also, function will compare values from Bit1 -> Bit8 and save when different on each row"""

    result = df[df['Signal_Name'] == signal_name]
    unique_rows = [result.iloc[0]]

    for _, row in result.iterrows():
        is_duplicate = False
        for unique_row in unique_rows:
            if (row['Bit_Name1':'Bit_Name8'] == unique_row['Bit_Name1':'Bit_Name8']).all():
                is_duplicate = True
                break
        if not is_duplicate:
            unique_rows.append(row)

    return pd.concat(unique_rows, axis=1).T


def save_to_excel(df, signal_name):
    """Save DataFrame to an Excel file with sheet name as signal name."""
    current_directory = os.path.dirname(os.path.abspath(__file__))
    filename = os.path.join(current_directory, "../Script/Public_results.xlsx")

    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    else:
        wb = openpyxl.load_workbook(filename)

    if signal_name not in wb.sheetnames:
        wb.create_sheet(title=signal_name)
    else:
        wb.remove(wb[signal_name])
        wb.create_sheet(title=signal_name)

    data = df.values.tolist()

    sheet = wb[signal_name]
    header = df.columns.tolist()
    sheet.append(header)

    for row in data:
        sheet.append(row)

    wb.save(filename)


if __name__ == "__main__":
    signal_names = get_user_input()
    df = read_csv()
    print(df)
    for signal_name in signal_names:
        result = search_signal(df, signal_name)
        print(f"Results for signal '{signal_name}':")
        print(result)
        save_to_excel(result, signal_name)
